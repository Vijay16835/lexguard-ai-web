#!/usr/bin/env python3
"""
LexGuard AI – CI Consolidated Report Generator
Merges Selenium, Appium, API, Security, and Performance results into:
  - Automation_Test_Report.xlsx
  - Passed_Test_Cases.xlsx
  - Failed_Test_Cases.xlsx
  - Execution_Summary.xlsx
  - ci_report.html
  - ci_summary.json
"""

import argparse
import json
import os
import glob
import xml.etree.ElementTree as ET
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("WARNING: openpyxl not available – Excel reports will be skipped.")

# ── Colors ────────────────────────────────────────────────────────────────────
DARK_HEADER   = "1E293B"
GREEN_COLOR   = "10B981"
RED_COLOR     = "EF4444"
AMBER_COLOR   = "F59E0B"
WHITE_COLOR   = "FFFFFF"
LIGHT_GREEN   = "DCFCE7"
LIGHT_RED     = "FEE2E2"
LIGHT_AMBER   = "FEF9C3"

def make_header_fill(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)

def make_font(bold=False, color=WHITE_COLOR, size=11):
    return Font(bold=bold, color=color, size=size, name="Segoe UI")

def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Parsers ───────────────────────────────────────────────────────────────────
def parse_mochawesome_json(path: str) -> list[dict]:
    records = []
    try:
        with open(path) as f:
            data = json.load(f)
        suites = data.get("results", [])
        def walk(suites):
            for suite in suites:
                for test in suite.get("tests", []):
                    status = "PASS" if test.get("pass") else ("FAIL" if test.get("fail") else "SKIP")
                    records.append({
                        "test_id": f"SEL_{len(records)+1:03d}",
                        "suite": suite.get("title", "Selenium"),
                        "test_name": test.get("fullTitle", test.get("title", "Unknown")),
                        "status": status,
                        "duration_ms": test.get("duration", 0),
                        "error": test.get("err", {}).get("message", "") if isinstance(test.get("err"), dict) else str(test.get("err", "")),
                        "framework": "Selenium/Mocha",
                    })
                walk(suite.get("suites", []))
        walk(suites)
    except Exception as e:
        print(f"  [WARN] Could not parse mochawesome JSON {path}: {e}")
    return records

def parse_junit_xml(path: str, framework: str = "API/Pytest") -> list[dict]:
    records = []
    try:
        tree = ET.parse(path)
        root = tree.getroot()
        suites = root.findall("testsuite") or [root]
        idx = 1
        for suite in suites:
            suite_name = suite.get("name", framework)
            for case in suite.findall("testcase"):
                fail_el = case.find("failure") or case.find("error")
                skip_el = case.find("skipped")
                if fail_el is not None:
                    status = "FAIL"
                    error = (fail_el.text or "").strip()[:300]
                elif skip_el is not None:
                    status = "SKIP"
                    error = ""
                else:
                    status = "PASS"
                    error = ""
                duration_ms = int(float(case.get("time", 0)) * 1000)
                records.append({
                    "test_id": f"API_{idx:03d}",
                    "suite": suite_name,
                    "test_name": f"{case.get('classname', '')}::{case.get('name', 'test')}",
                    "status": status,
                    "duration_ms": duration_ms,
                    "error": error,
                    "framework": framework,
                })
                idx += 1
    except Exception as e:
        print(f"  [WARN] Could not parse JUnit XML {path}: {e}")
    return records

def collect_all_results(artifacts_dir: str) -> list[dict]:
    all_records: list[dict] = []

    # 1. Selenium – Mochawesome JSON
    for p in glob.glob(f"{artifacts_dir}/**/LexGuard-Selenium*/**/*.json", recursive=True):
        if "full-report" in p or "Mochawesome" in p or "mochawesome" in p:
            print(f"  [Selenium] Parsing: {p}")
            all_records.extend(parse_mochawesome_json(p))

    # 2. API – JUnit XML
    for p in glob.glob(f"{artifacts_dir}/**/LexGuard-API*/**/*.xml", recursive=True):
        print(f"  [API] Parsing: {p}")
        all_records.extend(parse_junit_xml(p, "Backend API/Pytest"))

    # 3. Appium – Mochawesome JSON or allure-results
    for p in glob.glob(f"{artifacts_dir}/**/LexGuard-Appium*/**/*.json", recursive=True):
        if "report" in p.lower():
            print(f"  [Appium] Parsing: {p}")
            recs = parse_mochawesome_json(p)
            for r in recs:
                r["framework"] = "Appium/WebdriverIO"
                r["test_id"] = r["test_id"].replace("SEL_", "APP_")
            all_records.extend(recs)

    # 4. Appium Java – Surefire XML
    for p in glob.glob(f"{artifacts_dir}/**/surefire-reports/**/*.xml", recursive=True):
        print(f"  [AppiumJava] Parsing: {p}")
        all_records.extend(parse_junit_xml(p, "Appium Java/TestNG"))

    # Deduplicate test_id collisions
    seen_ids: dict[str, int] = {}
    for rec in all_records:
        base = rec["test_id"]
        if base in seen_ids:
            seen_ids[base] += 1
            rec["test_id"] = f"{base}_{seen_ids[base]}"
        else:
            seen_ids[base] = 0

    return all_records


# ── Excel Writers ─────────────────────────────────────────────────────────────
COLUMNS = [
    ("Test ID",       "test_id",      14),
    ("Suite / Module","suite",        28),
    ("Test Name",     "test_name",    55),
    ("Status",        "status",       10),
    ("Duration (ms)", "duration_ms",  14),
    ("Framework",     "framework",    22),
    ("Error Message", "error",        50),
]

def _write_sheet(ws, records: list[dict]):
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = make_font(bold=True, color=WHITE_COLOR, size=11)
        cell.fill = make_header_fill(DARK_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    STATUS_COLORS = {"PASS": (GREEN_COLOR, LIGHT_GREEN), "FAIL": (RED_COLOR, LIGHT_RED), "SKIP": (AMBER_COLOR, LIGHT_AMBER)}

    for row_idx, rec in enumerate(records, 2):
        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            val = rec.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(key == "error"))
            if key == "status":
                fg, bg = STATUS_COLORS.get(val, (DARK_HEADER, WHITE_COLOR))
                cell.font = Font(bold=True, color=fg, name="Segoe UI", size=10)
                cell.fill = PatternFill("solid", fgColor=bg)
            else:
                cell.font = Font(name="Segoe UI", size=10)

def write_automation_report(records, output_dir):
    if not EXCEL_AVAILABLE:
        return
    wb = openpyxl.Workbook()

    # Sheet 1 – All Tests
    ws_all = wb.active
    ws_all.title = "All Test Results"
    _write_sheet(ws_all, records)

    # Sheet 2 – Execution Summary
    ws_sum = wb.create_sheet("Execution Summary")
    total   = len(records)
    passed  = sum(1 for r in records if r["status"] == "PASS")
    failed  = sum(1 for r in records if r["status"] == "FAIL")
    skipped = sum(1 for r in records if r["status"] == "SKIP")
    pass_pct = f"{passed*100/max(total,1):.2f}%"
    total_dur_ms = sum(r.get("duration_ms", 0) for r in records)
    total_dur_s = total_dur_ms / 1000

    metrics = [
        ("Report Generated",       datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Project",                "LexGuard AI – Legal Document Analyzer"),
        ("Total Test Cases",       total),
        ("Passed",                 passed),
        ("Failed",                 failed),
        ("Skipped",                skipped),
        ("Pass Percentage",        pass_pct),
        ("Total Execution Time",   f"{total_dur_s:.1f}s ({total_dur_ms}ms)"),
        ("Frameworks Used",        "Selenium/Mocha · Appium/WDIO · Appium/TestNG · Pytest"),
    ]
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 45
    for ridx, (metric, val) in enumerate(metrics, 1):
        ws_sum.cell(ridx, 1, metric).font = Font(bold=True, name="Segoe UI", size=10, color=DARK_HEADER)
        ws_sum.cell(ridx, 2, val).font    = Font(name="Segoe UI", size=10)

    # Sheet 3 – Passed
    ws_pass = wb.create_sheet("Passed Tests")
    _write_sheet(ws_pass, [r for r in records if r["status"] == "PASS"])

    # Sheet 4 – Failed
    ws_fail = wb.create_sheet("Failed Tests")
    _write_sheet(ws_fail, [r for r in records if r["status"] == "FAIL"])

    path = os.path.join(output_dir, "Automation_Test_Report.xlsx")
    wb.save(path)
    print(f"  [Excel] Written: {path}")

def write_passed_report(records, output_dir):
    if not EXCEL_AVAILABLE:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Passed Test Cases"
    _write_sheet(ws, [r for r in records if r["status"] == "PASS"])
    path = os.path.join(output_dir, "Passed_Test_Cases.xlsx")
    wb.save(path)
    print(f"  [Excel] Written: {path}")

def write_failed_report(records, output_dir):
    if not EXCEL_AVAILABLE:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Failed Test Cases"
    _write_sheet(ws, [r for r in records if r["status"] == "FAIL"])
    path = os.path.join(output_dir, "Failed_Test_Cases.xlsx")
    wb.save(path)
    print(f"  [Excel] Written: {path}")

def write_summary_report(records, output_dir):
    if not EXCEL_AVAILABLE:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Execution Summary"

    total   = len(records)
    passed  = sum(1 for r in records if r["status"] == "PASS")
    failed  = sum(1 for r in records if r["status"] == "FAIL")
    skipped = sum(1 for r in records if r["status"] == "SKIP")
    total_dur = sum(r.get("duration_ms", 0) for r in records)

    # By framework
    frameworks = {}
    for r in records:
        fw = r.get("framework", "Unknown")
        if fw not in frameworks:
            frameworks[fw] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0}
        frameworks[fw]["total"] += 1
        frameworks[fw][r["status"].lower()] += 1

    headers = ["Metric", "Value"]
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22
    for cidx, h in enumerate(headers, 1):
        cell = ws.cell(1, cidx, h)
        cell.font = make_font(bold=True, color=WHITE_COLOR)
        cell.fill = make_header_fill(DARK_HEADER)
        cell.border = thin_border()

    rows = [
        ("Execution Date", datetime.now().strftime("%Y-%m-%d")),
        ("Total Tests",    total),
        ("Passed",         passed),
        ("Failed",         failed),
        ("Skipped",        skipped),
        ("Pass %",         f"{passed*100/max(total,1):.2f}%"),
        ("Fail %",         f"{failed*100/max(total,1):.2f}%"),
        ("Total Duration", f"{total_dur/1000:.1f}s"),
        ("", ""),
        ("Framework Breakdown", ""),
    ]
    for fw, stats in frameworks.items():
        rows.append((fw, f"Total={stats['total']} | Pass={stats['passed']} | Fail={stats['failed']}"))

    for ridx, (k, v) in enumerate(rows, 2):
        ws.cell(ridx, 1, k).font = Font(bold=bool(v == ""), name="Segoe UI", size=10)
        ws.cell(ridx, 2, str(v)).font = Font(name="Segoe UI", size=10)

    path = os.path.join(output_dir, "Execution_Summary.xlsx")
    wb.save(path)
    print(f"  [Excel] Written: {path}")

# ── HTML Report ───────────────────────────────────────────────────────────────
def write_html_report(records, output_dir):
    total   = len(records)
    passed  = sum(1 for r in records if r["status"] == "PASS")
    failed  = sum(1 for r in records if r["status"] == "FAIL")
    skipped = sum(1 for r in records if r["status"] == "SKIP")
    pass_pct = f"{passed*100/max(total,1):.1f}"

    rows_html = ""
    for r in records:
        sc = {"PASS": "#10b981", "FAIL": "#ef4444", "SKIP": "#f59e0b"}.get(r["status"], "#6b7280")
        err_snip = (r.get("error","") or "")[:120]
        rows_html += f"""
        <tr>
          <td class="mono">{r['test_id']}</td>
          <td>{r['suite']}</td>
          <td>{r['test_name'][:80]}</td>
          <td><span style="color:{sc};font-weight:700">{r['status']}</span></td>
          <td>{r['duration_ms']}ms</td>
          <td>{r['framework']}</td>
          <td class="error-cell" title="{err_snip}">{err_snip}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>LexGuard AI – CI Test Report</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
  :root{{--bg:#0f172a;--card:#1e293b;--border:#334155;--text:#f8fafc;--muted:#94a3b8}}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Inter,sans-serif;background:var(--bg);color:var(--text);padding:32px}}
  h1{{font-size:28px;background:linear-gradient(135deg,#38bdf8,#818cf8);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:8px}}
  .sub{{color:var(--muted);font-size:14px;margin-bottom:28px}}
  .kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:16px;margin-bottom:32px}}
  .kpi{{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:20px}}
  .kpi-label{{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px}}
  .kpi-value{{font-size:32px;font-weight:700}}
  table{{width:100%;border-collapse:collapse;background:var(--card);border-radius:12px;overflow:hidden;border:1px solid var(--border)}}
  th{{background:#111827;color:var(--muted);font-size:11px;font-weight:600;text-transform:uppercase;padding:12px 14px;text-align:left}}
  td{{padding:11px 14px;font-size:13px;border-bottom:1px solid var(--border)}}
  tr:hover{{background:rgba(255,255,255,.02)}}
  .mono{{font-family:monospace;font-size:12px}}
  .error-cell{{font-size:11px;color:#f87171;max-width:280px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
  .badge{{display:inline-block;padding:3px 10px;border-radius:20px;font-size:12px;font-weight:600;background:#1e293b;border:1px solid var(--border)}}
</style>
</head>
<body>
<h1>LexGuard AI – CI/CD Test Report</h1>
<div class="sub">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')} &nbsp;|&nbsp; Branch: GitHub Actions CI Pipeline</div>
<div class="kpi-grid">
  <div class="kpi"><div class="kpi-label">Total Tests</div><div class="kpi-value">{total}</div></div>
  <div class="kpi"><div class="kpi-label">Passed</div><div class="kpi-value" style="color:#10b981">{passed}</div></div>
  <div class="kpi"><div class="kpi-label">Failed</div><div class="kpi-value" style="color:#ef4444">{failed}</div></div>
  <div class="kpi"><div class="kpi-label">Skipped</div><div class="kpi-value" style="color:#f59e0b">{skipped}</div></div>
  <div class="kpi"><div class="kpi-label">Pass Rate</div><div class="kpi-value" style="color:#38bdf8">{pass_pct}%</div></div>
</div>
<table>
<thead><tr><th>Test ID</th><th>Suite</th><th>Test Name</th><th>Status</th><th>Duration</th><th>Framework</th><th>Error</th></tr></thead>
<tbody>{rows_html}</tbody>
</table>
</body>
</html>"""

    path = os.path.join(output_dir, "ci_report.html")
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  [HTML] Written: {path}")

# ── JSON Summary ──────────────────────────────────────────────────────────────
def write_json_summary(records, output_dir):
    total   = len(records)
    passed  = sum(1 for r in records if r["status"] == "PASS")
    failed  = sum(1 for r in records if r["status"] == "FAIL")
    skipped = sum(1 for r in records if r["status"] == "SKIP")
    dur_ms  = sum(r.get("duration_ms", 0) for r in records)

    summary = {
        "generated_at": datetime.now().isoformat(),
        "project": "LexGuard AI",
        "summary": {
            "total": total,
            "passed": passed,
            "failed": failed,
            "skipped": skipped,
            "pass_pct": round(passed * 100 / max(total, 1), 2),
            "total_duration_ms": dur_ms,
        },
        "tests": records,
    }
    path = os.path.join(output_dir, "ci_summary.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    print(f"  [JSON] Written: {path}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="LexGuard CI Report Generator")
    parser.add_argument("--artifacts-dir", default="downloaded-artifacts", help="Root dir with downloaded artifacts")
    parser.add_argument("--output-dir",    default="ci-reports",           help="Output directory for reports")
    parser.add_argument("--bandit",        default="",                     help="Path to bandit JSON (unused here)")
    parser.add_argument("--output",        default="",                     help="Alt output path (unused here)")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    print(f"\n🔍 Scanning artifacts in: {args.artifacts_dir}")
    records = collect_all_results(args.artifacts_dir)

    total   = len(records)
    passed  = sum(1 for r in records if r["status"] == "PASS")
    failed  = sum(1 for r in records if r["status"] == "FAIL")
    skipped = sum(1 for r in records if r["status"] == "SKIP")
    print(f"\n📊 Collected: {total} tests | {passed} passed | {failed} failed | {skipped} skipped")

    print("\n📝 Generating reports...")
    write_automation_report(records, args.output_dir)
    write_passed_report(records, args.output_dir)
    write_failed_report(records, args.output_dir)
    write_summary_report(records, args.output_dir)
    write_html_report(records, args.output_dir)
    write_json_summary(records, args.output_dir)

    print(f"\n✅ All reports written to: {args.output_dir}/")

if __name__ == "__main__":
    main()
